#!/usr/bin/env python3
"""
Gerador de indicadores a partir de Tabelas 2025/ (sem pandas).
Lê XLSX diretamente com openpyxl.
"""
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRO: openpyxl não instalado. Execute:")
    print("  pip install openpyxl")
    exit(1)

ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = ROOT / "Tabelas 2025"
OUT_DIR = ROOT / "data" / "indicadores" / "2025"
CATALOG = ROOT / "data" / "catalogo_2025.json"

MUNICIPALITY_CANDIDATES = [
    "Município", "Municipio", "MUNICÍPIO", "MUNICIPIO",
    "Municípios", "Município/UF", "Município (IBGE)",
    "Município - IBGE", "Nome do Município", "Municipality"
]

IGNORED_PREFIXES = ("~$",)
IGNORED_FILES = {"Thumbs.db"}

YEAR_RE = re.compile(r"(20\d{2})")


def slugify(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[\s_/]+", "-", text)
    text = re.sub(r"[^a-z0-9\-]", "", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def get_sheet_values(ws):
    """Extrai valores da worksheet, tratando merged cells."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([v for v in row])
    return rows


def pick_municipality_column(headers: List) -> Optional[int]:
    """Retorna o índice da coluna de município."""
    for cand in MUNICIPALITY_CANDIDATES:
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == cand.lower():
                return i
    return None


def pick_latest_year_column(headers: List, rows: List) -> Optional[Tuple[int, Optional[int]]]:
    """Retorna (índice_coluna, ano) da coluna numérica com ano mais recente."""
    candidates: List[Tuple[int, int]] = []
    for i, h in enumerate(headers):
        if not h:
            continue
        m = YEAR_RE.search(str(h))
        if m:
            year = int(m.group(1))
            # verifica se coluna tem valores numéricos
            has_nums = False
            for row in rows[1:]:  # skip header
                if i < len(row) and row[i] is not None:
                    try:
                        float(row[i])
                        has_nums = True
                        break
                    except (ValueError, TypeError):
                        pass
            if has_nums:
                candidates.append((i, year))
    if candidates:
        i, year = max(candidates, key=lambda x: x[1])
        return i, year
    # fallback: primeira coluna numérica
    for i, h in enumerate(headers):
        for row in rows[1:]:
            if i < len(row) and row[i] is not None:
                try:
                    float(row[i])
                    return i, None
                except (ValueError, TypeError):
                    pass
    return None


def infer_unit(colname: str) -> str:
    c = str(colname).lower() if colname else ""
    if "r$" in c or "preço" in c or "valor" in c:
        return "R$"
    if "percent" in c or "%" in c:
        return "%"
    if "idh" in c:
        return ""
    if "hab/km" in c:
        return "hab/km²"
    if "hab" in c or "popula" in c:
        return "hab"
    if "km2" in c or "km²" in c:
        return "km²"
    return ""


def clean_label(filename: str, colname: str, year: Optional[int]) -> str:
    base = Path(filename).stem
    base = re.sub(r"^tab\s*\d+[\w.\s-]*\s*", "", base, flags=re.IGNORECASE)
    base = base.replace("_", " ").strip()
    label = base
    if year:
        label = f"{label} ({year})"
    return re.sub(r"\s+", " ", label).strip()


def process_file(xlsx_path: Path, category_key: str) -> List[Dict]:
    indicators = []
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  ⚠ Erro ao ler {xlsx_path.name}: {e}")
        return indicators

    rows = get_sheet_values(ws)
    if len(rows) < 2:
        return indicators

    headers = rows[0]
    data_rows = rows[1:]

    mun_col_idx = pick_municipality_column(headers)
    if mun_col_idx is None:
        return indicators

    result = pick_latest_year_column(headers, rows)
    if not result:
        return indicators

    val_col_idx, year = result
    unit = infer_unit(headers[val_col_idx] if val_col_idx < len(headers) else "")
    label = clean_label(xlsx_path.name, headers[val_col_idx] if val_col_idx < len(headers) else "", year)
    slug = slugify(f"{category_key}-{xlsx_path.stem}-{year or 'latest'}")

    # Monta mapping municipio -> valor (float)
    mapping = {}
    for row in data_rows:
        if mun_col_idx >= len(row) or val_col_idx >= len(row):
            continue
        mun = row[mun_col_idx]
        val = row[val_col_idx]
        if mun is None or val is None:
            continue
        mun_str = str(mun).strip()
        try:
            val_float = float(val)
        except (ValueError, TypeError):
            continue
        mapping[mun_str] = val_float

    if not mapping:
        return indicators

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUT_DIR / f"{slug}.json"
    payload = {
        "category": category_key,
        "label": label,
        "unit": unit,
        "year": year,
        "source_file": str(xlsx_path.relative_to(ROOT)),
        "data": mapping,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"  ✓ {label} ({len(mapping)} municípios) → {out_path.name}")

    indicators.append({
        "slug": slug,
        "label": label,
        "unit": unit,
        "year": year,
        "path": str(out_path.relative_to(ROOT)).replace("\\", "/"),
    })
    return indicators


def main():
    if not SRC_DIR.exists():
        print(f"❌ Pasta não encontrada: {SRC_DIR}")
        return

    categories = [
        ("demografia", "1. Demografia"),
        ("economia", "2. Economia"),
        ("infraestrutura", "3. Infraestrutura"),
        ("meio-ambiente", "4. Meio Ambiente"),
        ("social", "5. Social"),
    ]

    catalog: Dict[str, List[Dict]] = {k: [] for k, _ in categories}
    total = 0

    for key, folder_name in categories:
        folder = SRC_DIR / folder_name
        if not folder.exists():
            print(f"⚠ Pasta não encontrada: {folder_name}")
            continue

        print(f"\n📂 {folder_name}")
        
        # Arquivos diretos
        for path in sorted(folder.glob("*.xlsx")):
            if path.name.startswith(IGNORED_PREFIXES) or path.name in IGNORED_FILES:
                continue
            inds = process_file(path, key)
            catalog[key].extend(inds)
            total += len(inds)

        # Subpastas
        for subfolder in sorted(folder.iterdir()):
            if subfolder.is_dir():
                print(f"   📁 {subfolder.name}")
                for path in sorted(subfolder.glob("*.xlsx")):
                    if path.name.startswith(IGNORED_PREFIXES) or path.name in IGNORED_FILES:
                        continue
                    inds = process_file(path, key)
                    catalog[key].extend(inds)
                    total += len(inds)

    CATALOG.parent.mkdir(parents=True, exist_ok=True)
    with open(CATALOG, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Catálogo gerado com {total} indicadores: {CATALOG}")
    print(f"📊 JSONs salvos em: {OUT_DIR}")


if __name__ == "__main__":
    main()
